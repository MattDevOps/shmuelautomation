from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def _payload(**overrides: object) -> dict[str, object]:
    base: dict[str, object] = {
        "type": "rent",
        "price": "8500.00",
        "rooms": "3.5",
        "size_sqm": 80,
        "neighborhood": "Baka",
        "owner_name": "Yossi",
        "owner_phone": "+972500000000",
        "broker_fee_status": "yes",
        "description": "Bright 3.5-room flat",
    }
    base.update(overrides)
    return base


def test_export_returns_xlsx_with_correct_headers(client: TestClient) -> None:
    client.post("/properties", json=_payload())

    r = client.get("/properties/export")
    assert r.status_code == 200
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        r.headers["content-disposition"]
        == f'attachment; filename="properties-{date.today().isoformat()}.xlsx"'
    )


def test_export_workbook_has_header_row_and_property_rows(
    client: TestClient,
) -> None:
    client.post("/properties", json=_payload(neighborhood="Baka"))
    client.post("/properties", json=_payload(neighborhood="Katamon", price="9500"))

    r = client.get("/properties/export")
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    assert ws.title == "Properties"

    headers = [c.value for c in ws[1]]
    assert headers[:5] == ["ID", "Type", "Status", "Price", "Currency"]
    assert "Owner phone" in headers
    assert "Notes" in headers
    assert "Broker fee" in headers

    neighborhood_col = headers.index("Neighborhood") + 1
    body_cells = [
        ws.cell(row=row, column=neighborhood_col).value
        for row in range(2, ws.max_row + 1)
    ]
    assert sorted(body_cells) == ["Baka", "Katamon"]


def test_export_includes_sensitive_internal_fields(client: TestClient) -> None:
    client.post(
        "/properties",
        json=_payload(owner_phone="+972500000000", notes="haggle ok at 8200"),
    )

    r = client.get("/properties/export")
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws is not None

    headers = [c.value for c in ws[1]]
    phone_col = headers.index("Owner phone") + 1
    notes_col = headers.index("Notes") + 1
    assert ws.cell(row=2, column=phone_col).value == "+972500000000"
    assert ws.cell(row=2, column=notes_col).value == "haggle ok at 8200"


def test_export_with_no_properties_returns_header_only_workbook(
    client: TestClient,
) -> None:
    r = client.get("/properties/export")
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    assert ws.max_row == 1
